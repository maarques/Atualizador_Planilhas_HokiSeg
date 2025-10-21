import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
import math
import calendar

from parsers import porto_seguro
from parsers import amil

def processar_planilha(pdf_file, planilha_file, seguradora_escolhida): # <-- MUDANÇA: Novo argumento
    """
    Função principal de processamento.
    Recebe a seguradora, chama o parser correto, e atualiza a planilha.
    """
    sheet_name_to_update = "Comissão" 
    dir_name = os.path.dirname(planilha_file)
    mes_atual = calendar.month_name[datetime.now().month].lower()
    ano_atual = datetime.now().year
    novo_arquivo = os.path.join(dir_name, f"Planilha_Financeira_{mes_atual}_{ano_atual}_Automatizada.xlsx")

    # ESCOLHER O PARSER E LER OS DADOS
    print(f"ETAPA 1: Selecionando parser para '{seguradora_escolhida}'...")
    
    dados_extraidos = []
    
    if seguradora_escolhida == "Porto Seguro":
        dados_extraidos = porto_seguro.parse(pdf_file)
    elif seguradora_escolhida == "Amil":
        dados_extraidos = amil.parse(pdf_file)
    else:
        print(f"❌ ERRO: Nenhum parser foi encontrado para '{seguradora_escolhida}'.")
        return

    if not dados_extraidos:
        print("⚠️ Nenhum dado foi retornado pelo parser. Encerrando.")
        return

    print(f"Dados brutos recebidos do parser: {len(dados_extraidos)} linhas.")
    
    # CRIAR DATAFRAME E LIMPAR DADOS
    print("ETAPA 2: Limpando e convertendo dados...")
    df = pd.DataFrame(dados_extraidos, columns=[
        'Historico', 'Marca', 'Apolice', 'Parcela', 'Data_Pag', 
        'Premio', 'Taxa', 'Comissao'
    ])

    try:
        df['Premio_float'] = pd.to_numeric(df['Premio'].str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
        df['Comissao_float'] = pd.to_numeric(df['Comissao'].str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
        df['Taxa_float'] = pd.to_numeric(df['Taxa'].str.replace(",", ".", regex=False)) / 100.0
        df['Parcela_int'] = pd.to_numeric(df['Parcela'])
        df['Data_obj'] = pd.to_datetime(df['Data_Pag'])
    except ValueError as e:
        print(f"❌ ERRO: Falha ao converter dados. Verifique o formato do PDF. Erro: {e}")
        raise 
    
    df = df.dropna(subset=['Premio_float', 'Comissao_float', 'Taxa_float', 'Parcela_int', 'Data_obj'])

    # AGRUPAR E SOMAR
    print("ETAPA 3: Agrupando e somando dados...")
    group_keys = ['Historico', 'Apolice', 'Parcela_int']

    df_agrupado = df.groupby(group_keys, as_index=False).agg(
        Valor_Total=('Premio_float', 'sum'),
        Comissao_Total=('Comissao_float', 'sum'),
        Marca=('Marca', 'first'),
        Data_obj=('Data_obj', 'first'),
        Taxa_float=('Taxa_float', 'max') 
    )

    df_agrupado['Comissao_Total'] = df_agrupado['Comissao_Total'].apply(lambda x: math.ceil(x * 100) / 100)
    
    print(f"Dados lidos e agrupados. {len(df_agrupado)} linhas únicas serão inseridas.")

    # PREPARAR DADOS (ADICIONAR COLUNAS FIXAS)
    print("ETAPA 4: Preparando colunas finais...")
    df_agrupado = df_agrupado.rename(columns={
        'Historico': 'Cliente', 'Apolice': 'Apólice', 'Parcela_int': 'Parcela',
        'Data_obj': 'Dt. Pagamento', 'Taxa_float': 'Porcentagem',
        'Valor_Total': 'Valor', 'Comissao_Total': 'Valor Comissão'
    })

    df_agrupado['Seguradora'] = df_agrupado['Marca'].apply(
        lambda x: "Porto Seguro" if "Porto" in x else x
    )
    df_agrupado['Parcela a receber'] = 12
    df_agrupado['Situação'] = 'Pago'
    df_agrupado['Tipo'] = 'Vida Presente'
    df_agrupado['Colaborador'] = 'Calina'

    # INSERIR NAS COLUNAS CORRETAS
    print(f"ETAPA 5: Abrindo a planilha '{os.path.basename(planilha_file)}'...")
    try:
        wb = load_workbook(planilha_file)
    except FileNotFoundError:
        print(f"❌ ERRO: Arquivo '{planilha_file}' não encontrado. Verifique o nome.")
        raise

    try:
        ws = wb[sheet_name_to_update]
    except KeyError:
        print(f"❌ ERRO: A aba '{sheet_name_to_update}' não foi encontrada no arquivo Excel.")
        print(f"Abas disponíveis: {wb.sheetnames}")
        raise

    first_empty_row = 5 
    while ws.cell(row=first_empty_row, column=2).value is not None: 
        first_empty_row += 1

    print(f"Inserindo {len(df_agrupado)} novas linhas a partir da linha {first_empty_row}...")

    colunas_para_escrever = {
        2: 'Apólice', 3: 'Valor', 5: 'Cliente', 6: 'Seguradora',
        7: 'Parcela a receber', 8: 'Parcela', 9: 'Dt. Pagamento',
        10: 'Valor Comissão', 11: 'Porcentagem', 12: 'Situação',
        13: 'Tipo', 14: 'Colaborador'
    }

    for _, row in df_agrupado.iterrows():
        for col_idx, col_name in colunas_para_escrever.items():
            ws.cell(row=first_empty_row, column=col_idx, value=row[col_name])
        
        ws.cell(row=first_empty_row, column=3).number_format = '"R$" #,##0.00'
        ws.cell(row=first_empty_row, column=9).number_format = 'DD/MM/YYYY'
        ws.cell(row=first_empty_row, column=10).number_format = '"R$" #,##0.00'
        ws.cell(row=first_empty_row, column=11).number_format = '0.00%'
        first_empty_row += 1 

    print(f"Salvando como '{os.path.basename(novo_arquivo)}'...")
    wb.save(novo_arquivo)

    # MOSTRAR RESULTADOS
    print(f"\n✅ Planilha atualizada com sucesso: {novo_arquivo}")
    print(f"➕ Foram adicionadas {len(df_agrupado)} linhas (agrupadas e somadas):\n")

    print(f"{'Apólice':<15} | {'Cliente':<30} | {'Parc.':>5} | {'Valor':>10} | {'V. Comissão':>10} | {'Taxa':>6}")
    print("-" * 85)
    for _, row in df_agrupado.iterrows():
        print(f"{row['Apólice']:<15} | {row['Cliente']:<30} | {row['Parcela']:>5} | {row['Valor']:>10.2f} | {row['Valor Comissão']:>10.2f} | {row['Porcentagem']:>6.2%}")
